from openpyxl import Workbook
import os

class XLSync:
    def __init__(self, path, name):
        self.path = path
        self.name = name

    def execute(self):
        wb = Workbook()
        ws = wb.active
        count = 0
        for dir, subdir, file in os.walk(self.path):
            index=0
            while count< len(file):
                ws.cell(row=count+1, column = 1, value=file[index])
                count+=1
                index+=1

        wb.save("{0}.xlsx".format(self.name))


XLSync("C://Users//Shane_Programming//Desktop", "test_upload").execute()
# wb = Workbook()
# ws = wb.active
# x = ["Shane", "John", "Keith", "Alex"]
#
# c=0
# while c < len(x):
#     ws.cell(row=c+1, column= 1, value=x[c])
#     c+=1
#
# wb.save("test_upload.xlsx")
